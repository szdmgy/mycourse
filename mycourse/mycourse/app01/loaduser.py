import openpyxl

def load_user_list():
    file = 'dmc.xlsx'
    wb = openpyxl.load_workbook(file)
    ws = wb[wb.sheetnames[0]]
    # print(ws.title)
    # print(ws.max_row  )
    row = ws.max_row
    col = ws.max_column
    # print(row,col)
    users = []
    for i in range(10,row):
        number, name,sex = ws.cell(i, 2).value, ws.cell(i, 3).value,ws.cell(i, 4).value
        if number and name and sex:
            # print(number,'szu'+number[len(number)-6:], name,sex)
            users.append((number,name,sex))
        else:
            break
    return users



if __name__ == '__main__':
    # print('hello')
    for each in load_user_list():
        print(each[0][-6:])