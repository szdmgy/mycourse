import os
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from collections import OrderedDict
from .upload_data import extract_course_data


@csrf_exempt
def import_course(request):
    """
    处理多Excel文件上传视图函数
    """
    if not request.user.is_superuser:
        return HttpResponse("警告！您不是管理员！无法进入此界面！")

    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': '仅支持POST方法'
        }, status=405)

    # 确保files在request.FILES中
    if not request.FILES:
        return JsonResponse({
            'success': False,
            'error': '未收到任何文件'
        }, status=400)

    results = []
    processed_files = 0
    sheet_count = 0
    total_rows = 0

    try:
        # 遍历所有上传的文件
        for file_key in request.FILES:
            file_obj = request.FILES[file_key]
            file_data = OrderedDict()
            file_data['filename'] = file_obj.name

            # 检查文件扩展名
            if not file_obj.name.lower().endswith(('.xlsx', '.xls')):
                file_data['error'] = '无效的文件格式'
                results.append(file_data)
                continue

            # extract_course_data(file_obj)
            # # 处理Excel文件
            # wb = load_workbook(file_obj, read_only=True, data_only=True)
            # sheet_count = len(wb.sheetnames)
            # total_rows = 0
            #
            # # 遍历所有工作表
            # for sheet_name in wb.sheetnames:
            #     sheet = wb[sheet_name]
            #     rows = []
            #
            #     # 逐行读取数据（跳过空行）
            #     for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            #         # 跳过全空行
            #         if all(cell is None for cell in row):
            #             continue
            #
            #         # 添加到结果
            #         rows.append(list(row))
            #         total_rows += 1
            #
            # 保存结果
            sheet_count += 1
            total_rows += 1
            file_data['sheet_count'] = sheet_count
            file_data['total_rows'] = total_rows
            results.append(file_data)
            processed_files += 1

        # 返回成功响应
        return JsonResponse({
            'success': True,
            'file_count': processed_files,
            'results': results
        })

    except Exception as e:
        # 捕获并返回异常信息
        return JsonResponse({
            'success': False,
            'error': f'处理过程中出错: {str(e)}'
        }, status=500)
