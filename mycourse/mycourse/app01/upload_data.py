import logging
import os
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.db import transaction
from app01 import models

logger = logging.getLogger('app01')

_hash_workers = min(os.cpu_count() or 4, 16)


def _bulk_make_passwords(plain_list):
    """多线程并行哈希密码，利用多核 CPU 加速"""
    if len(plain_list) <= 2:
        return [make_password(p) for p in plain_list]
    with ThreadPoolExecutor(max_workers=_hash_workers) as pool:
        return list(pool.map(make_password, plain_list))


# ═══════════════════════ 旧版一步式接口（保留兼容） ═══════════════════════

def extract_import_data(upload_file, file_format):
    if file_format == 'course':
        return extract_course_data(upload_file)
    elif file_format == 'task':
        return {'error': '作业导入请从课程详情页的"导入实验"功能操作'}
    elif file_format == 'student':
        return extract_student_data(upload_file)
    elif file_format == 'teacher':
        return extract_teacher_data(upload_file)
    elif file_format == 'user':
        return {'success': '用户数据文件解析成功'}
    else:
        return {'error': f'{upload_file}文件解析失败'}


def extract_course_data(upload_file):
    try:
        parsed = parse_course_excel(upload_file)
        if 'error' in parsed:
            return parsed
        write_course_data(parsed)
        return {'success': '课程数据文件解析成功'}
    except Exception as e:
        logger.exception("课程数据解析失败")
        return {'error': f'课程数据文件解析失败：{str(e)}'}


def parse_task_excel(upload_file, course):
    """纯解析作业 Excel，返回预览数据（不写入数据库）。
    Excel 列：A标题  B内容  C文件类型(可选)  D显示(Y/N, 可选)
    同时检测数据库重复和 Excel 文件内重复。
    """
    try:
        wb = openpyxl.load_workbook(upload_file)
        ws = wb[wb.sheetnames[0]]
        existing_titles = set(
            models.Task.objects.filter(courseBelongTo=course)
            .values_list('title', flat=True)
        )
        tasks = []
        seen_titles = set()
        for row in range(2, ws.max_row + 1):
            title = ws.cell(row, 1).value
            content = ws.cell(row, 2).value
            if not title or not content:
                break
            title = str(title).strip()
            content = str(content).strip()

            file_type = ws.cell(row, 3).value or '*'
            file_type = str(file_type).strip()
            display_val = ws.cell(row, 4).value or ''
            display = False if str(display_val).strip().upper() == 'N' else True

            if title in existing_titles:
                dup_reason = 'db'
            elif title in seen_titles:
                dup_reason = 'file'
            else:
                dup_reason = None

            seen_titles.add(title)
            tasks.append({
                'title': title, 'content': content,
                'fileType': file_type, 'display': display,
                'duplicate': dup_reason is not None,
                'dup_reason': dup_reason,
            })

        if not tasks:
            return {'error': 'Excel 中未解析到有效作业数据'}
        return {'tasks': tasks}
    except Exception as e:
        logger.exception("作业 Excel 解析失败")
        return {'error': f'Excel 解析失败：{str(e)}'}


@transaction.atomic
def write_task_import(tasks_data, course):
    """批量将预览确认后的作业数据写入数据库（仅写入非重复项）。"""
    from datetime import date, timedelta
    default_deadline = date.today() + timedelta(days=120)
    new_tasks = [
        models.Task(
            title=t['title'], content=t['content'],
            courseBelongTo=course,
            deadline=default_deadline, display=t['display'],
            fileType=t.get('fileType', '*'),
        )
        for t in tasks_data if not t.get('duplicate')
    ]
    if new_tasks:
        models.Task.objects.bulk_create(new_tasks)
    return {'success': f'导入完成：新增 {len(new_tasks)} 个作业'}


def extract_student_data(upload_file):
    try:
        wb = openpyxl.load_workbook(upload_file)
        ws = wb[wb.sheetnames[0]]
        rows = ws.max_row
        student_list = []
        for row in range(1, rows + 1):
            number, name, sex = ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 3).value
            if number and name and sex:
                student_list.append([str(number), name, sex])
            else:
                break
        write_student_users(student_list)
        return {'success': '学生数据文件解析成功'}
    except Exception as e:
        logger.exception("学生数据解析失败")
        return {'error': f'学生数据文件解析失败：{str(e)}'}


def extract_teacher_data(upload_file):
    try:
        wb = openpyxl.load_workbook(upload_file)
        ws = wb[wb.sheetnames[0]]
        rows = ws.max_row
        teacher_list = []
        for row in range(1, rows + 1):
            number, name, sex = ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 3).value
            if number and name and sex:
                teacher_list.append([str(number), name, sex])
            else:
                break
        write_teacher_users(teacher_list)
        return {'success': '老师数据文件解析成功'}
    except Exception as e:
        logger.exception("教师数据解析失败")
        return {'error': f'老师数据文件解析失败：{str(e)}'}


# ═══════════════════════ 第一层：纯解析（不访问 DB） ═══════════════════════

def parse_course_excel(upload_file):
    """解析课程 Excel，返回结构化数据（不写库、不查库）"""
    try:
        wb = openpyxl.load_workbook(upload_file)
        ws = wb[wb.sheetnames[0]]

        def cell_str(row, col):
            v = ws.cell(row, col).value
            return str(v).strip() if v is not None else ''

        course_term = cell_str(3, 1)
        course_number = cell_str(5, 3)
        course_name = cell_str(5, 18)
        class_number = cell_str(5, 6)
        teachers_str = cell_str(6, 11)

        errors = []
        if not course_term:
            errors.append('学期（A3）为空')
        if not course_number:
            errors.append('课程编号（C5）为空')
        if not course_name:
            errors.append('课程名（R5）为空')
        if not class_number:
            errors.append('班号（F5）为空')

        teacher_names = [t.strip() for t in teachers_str.split(',') if t.strip()] if teachers_str else []

        students = []
        for row in range(10, ws.max_row + 1):
            number = ws.cell(row, 2).value
            name = ws.cell(row, 3).value
            sex = ws.cell(row, 4).value
            if not number and not name:
                break
            if number and name:
                students.append({
                    'number': str(number).strip(),
                    'name': str(name).strip(),
                    'gender': str(sex).strip() if sex else '男',
                })
            else:
                errors.append(f'第{row}行数据不完整：学号={number}, 姓名={name}')

        if errors:
            return {'error': '；'.join(errors)}

        return {
            'courseTerm': course_term,
            'courseNumber': course_number,
            'courseName': course_name,
            'classNumber': class_number,
            'teachers': teacher_names,
            'students': students,
        }
    except Exception as e:
        logger.exception("课程 Excel 解析异常")
        return {'error': f'文件解析失败：{str(e)}'}


# ═══════════════════════ 第二层：预览（读 DB 标注状态，不写 DB） ═══════════════════════

def preview_course_import(parsed):
    """对比 DB 标注每条数据的状态，检测名单变动，不写库"""
    result = {
        'course': {
            'courseTerm': parsed['courseTerm'],
            'courseNumber': parsed['courseNumber'],
            'courseName': parsed['courseName'],
            'classNumber': parsed['classNumber'],
        },
        'teachers': [],
        'students': [],
        'removed_students': [],
        'summary': {},
    }

    existing = models.Course.objects.filter(
        courseTerm=parsed['courseTerm'],
        courseNumber=parsed['courseNumber'],
        classNumber=parsed['classNumber'],
    ).first()
    result['course']['exists'] = existing is not None

    teacher_names = parsed['teachers']
    teacher_profiles = {
        p.name: p.user.username
        for p in models.UserProfile.objects.filter(
            name__in=teacher_names, type='T'
        ).select_related('user')
    }
    for tname in teacher_names:
        if tname in teacher_profiles:
            result['teachers'].append({
                'name': tname, 'number': teacher_profiles[tname], 'status': '已存在',
            })
        else:
            result['teachers'].append({
                'name': tname, 'number': '—', 'status': '仅关联名称(无账号)',
            })

    excel_numbers = {stu['number'] for stu in parsed['students']}
    all_numbers = list(excel_numbers)
    existing_profiles = {
        p.user.username: p.name
        for p in models.UserProfile.objects.filter(
            user__username__in=all_numbers
        ).select_related('user')
    }

    new_count = 0
    exist_count = 0
    error_count = 0
    for stu in parsed['students']:
        if stu['number'] in existing_profiles:
            db_name = existing_profiles[stu['number']]
            if db_name != stu['name']:
                result['students'].append({
                    **stu, 'status': f'已存在(姓名不一致: 库中={db_name})',
                    'conflict': True,
                })
                error_count += 1
            else:
                result['students'].append({**stu, 'status': '已存在', 'conflict': False})
                exist_count += 1
        else:
            result['students'].append({**stu, 'status': '新建账号', 'conflict': False})
            new_count += 1

    removed_count = 0
    if existing:
        for profile in existing.members.filter(type='S').select_related('user'):
            if profile.user.username not in excel_numbers:
                result['removed_students'].append({
                    'number': profile.user.username,
                    'name': profile.name,
                })
                removed_count += 1

    has_changes = new_count > 0 or removed_count > 0
    if existing:
        if has_changes:
            result['course']['status'] = '已存在'
            result['course']['action'] = 'update'
        else:
            result['course']['status'] = '已存在，名单无变化'
            result['course']['action'] = 'none'
    else:
        result['course']['status'] = '新建'
        result['course']['action'] = 'create'

    result['summary'] = {
        'student_new': new_count,
        'student_exist': exist_count,
        'student_removed': removed_count,
        'student_error': error_count,
        'student_total': len(parsed['students']),
        'teacher_count': len(parsed['teachers']),
        'course_status': result['course']['status'],
        'action': result['course']['action'],
        'has_changes': has_changes,
    }

    return result


# ═══════════════════════ 第三层：写入 DB ═══════════════════════

@transaction.atomic
def write_course_data(parsed):
    """将解析后的课程数据写入数据库，支持新增和更新名单（含移除退出学生）"""
    course_obj, created = models.Course.objects.get_or_create(
        courseTerm=parsed['courseTerm'],
        courseNumber=parsed['courseNumber'],
        classNumber=parsed['classNumber'],
        defaults={
            'courseName': parsed['courseName'],
            'teachers': ','.join(parsed['teachers']) if isinstance(parsed['teachers'], list) else parsed.get('teachers', ''),
        }
    )

    if isinstance(parsed['students'], list) and parsed['students']:
        if isinstance(parsed['students'][0], dict):
            student_list = [[s['number'], s['name'], s.get('gender', '男')] for s in parsed['students']]
        else:
            student_list = parsed['students']
    else:
        student_list = parsed.get('students', [])

    write_student_users(student_list)

    numbers = [stu[0] if isinstance(stu, list) else stu['number'] for stu in student_list]
    student_profiles = list(models.UserProfile.objects.filter(user__username__in=numbers))
    course_obj.members.add(*student_profiles)

    if not created:
        to_remove = list(
            course_obj.members.filter(type='S').exclude(user__username__in=numbers)
        )
        if to_remove:
            course_obj.members.remove(*to_remove)

    teachers = parsed['teachers']
    if isinstance(teachers, str):
        teachers = [t.strip() for t in teachers.split(',') if t.strip()]
    teacher_profiles = list(models.UserProfile.objects.filter(name__in=teachers, type='T'))
    if teacher_profiles:
        course_obj.members.add(*teacher_profiles)

    return course_obj


@transaction.atomic
def write_student_users(student_list):
    """批量创建学生账号（已存在则跳过）"""
    normalized = []
    for stu in student_list:
        if isinstance(stu, dict):
            normalized.append((str(stu['number']), stu['name'], stu.get('gender', '男')))
        else:
            normalized.append((str(stu[0]), stu[1], stu[2] if len(stu) > 2 else '男'))

    all_numbers = [n for n, _, _ in normalized]
    existing = set(User.objects.filter(username__in=all_numbers).values_list('username', flat=True))

    data_map = {n: (name, gender) for n, name, gender in normalized if n not in existing}
    if not data_map:
        return

    numbers = list(data_map.keys())
    hashed = _bulk_make_passwords(['szu' + n[-6:] for n in numbers])
    new_users = [User(username=n, password=h) for n, h in zip(numbers, hashed)]
    User.objects.bulk_create(new_users)

    created_users = User.objects.filter(username__in=numbers)
    new_profiles = [
        models.UserProfile(
            user=u, name=data_map[u.username][0], type='S',
            gender='M' if data_map[u.username][1] == '男' else 'F',
        )
        for u in created_users
    ]
    models.UserProfile.objects.bulk_create(new_profiles)


def parse_teacher_excel(upload_file):
    """纯解析教师 Excel（不访问 DB）。格式：A工号 B姓名 C性别"""
    try:
        wb = openpyxl.load_workbook(upload_file)
        ws = wb[wb.sheetnames[0]]
        teachers = []
        for row in range(1, ws.max_row + 1):
            number = ws.cell(row, 1).value
            name = ws.cell(row, 2).value
            sex = ws.cell(row, 3).value
            if not number or not name:
                break
            teachers.append({
                'number': str(number).strip(),
                'name': str(name).strip(),
                'gender': str(sex).strip() if sex else '男',
            })
        if not teachers:
            return {'error': 'Excel 中未解析到有效教师数据'}
        return {'teachers': teachers}
    except Exception as e:
        logger.exception("教师 Excel 解析失败")
        return {'error': f'文件解析失败：{str(e)}'}


def preview_teacher_import(parsed):
    """对比 DB 标注每个教师的状态"""
    result = {'teachers': [], 'summary': {}}
    all_numbers = [t['number'] for t in parsed['teachers']]
    existing_profiles = {
        p.user.username: p.name
        for p in models.UserProfile.objects.filter(
            user__username__in=all_numbers
        ).select_related('user')
    }

    new_count = 0
    exist_count = 0
    for t in parsed['teachers']:
        if t['number'] in existing_profiles:
            db_name = existing_profiles[t['number']]
            if db_name != t['name']:
                result['teachers'].append({
                    **t, 'status': f'已存在(姓名不一致: 库中={db_name})', 'conflict': True,
                })
            else:
                result['teachers'].append({**t, 'status': '已存在', 'conflict': False})
            exist_count += 1
        else:
            result['teachers'].append({**t, 'status': '新建账号', 'conflict': False})
            new_count += 1
    result['summary'] = {
        'total': len(parsed['teachers']),
        'new': new_count,
        'exist': exist_count,
    }
    return result


@transaction.atomic
def write_teacher_users(teacher_list):
    """批量创建教师账号（已存在则跳过）"""
    normalized = []
    for teacher in teacher_list:
        if isinstance(teacher, dict):
            normalized.append((str(teacher['number']), teacher['name'], teacher.get('gender', '男')))
        else:
            normalized.append((str(teacher[0]), teacher[1], teacher[2] if len(teacher) > 2 else '男'))

    all_numbers = [n for n, _, _ in normalized]
    existing = set(User.objects.filter(username__in=all_numbers).values_list('username', flat=True))

    data_map = {n: (name, gender) for n, name, gender in normalized if n not in existing}
    if not data_map:
        return

    numbers = list(data_map.keys())
    hashed = _bulk_make_passwords(['szu' + n for n in numbers])
    new_users = [User(username=n, password=h) for n, h in zip(numbers, hashed)]
    User.objects.bulk_create(new_users)

    created_users = User.objects.filter(username__in=numbers)
    new_profiles = [
        models.UserProfile(
            user=u, name=data_map[u.username][0], type='T',
            gender='M' if data_map[u.username][1] == '男' else 'F',
        )
        for u in created_users
    ]
    models.UserProfile.objects.bulk_create(new_profiles)


def write_task_data(tasks):
    """已废弃，保留空壳兼容旧调用"""
    logger.warning("write_task_data 已废弃，请使用 parse_task_excel + write_task_import")
